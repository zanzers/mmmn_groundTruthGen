from openpyxl import Workbook, load_workbook
from pathlib import Path



def appendData_xl(exel_path: str, label: str, result: dict, image_id: int) -> None:
    
    headers = ["No", "Remarks"] + list(result.keys())
    values = [image_id, label] + list(result.values())

    try:
        
        if not Path(exel_path).exists():
            create_exel(exel_path, headers, row=values)
            return True
        else:
            return update_exel(exel_path, headers, values)
        
    except Exception as e:
        print(f"[ERROR] failed to save {e}") 
        return False


def create_exel(exel_path: str, headers: str, row: list = None) -> None:

    wb = Workbook()
    ws = wb.active
    ws.title = "mmmn_GroundTruth_datasets"

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    if row:
        for col, val in enumerate(row, start=1):
            ws.cell(row=2, column=col).value = val
    wb.save(exel_path)
    print(f"[ALERT] Succesfull creating new Exel file: {exel_path}")


def update_exel(exel_path: str, headers: str, values: list) -> bool:


    try:
        file = Path(exel_path)

        if not file.exists():
            print(f"[ERROR] Exel file not found!")
            return False

        wb = load_workbook(exel_path)

        if not wb.sheetnames:
            print(f"[ERROR] No valid worksheets")
            wb.save(exel_path)
        else:
            ws = wb.active

        if ws is None:
            print(f"[ERROR] No valid Worksheet")
            return False
        
        next_row = ws.max_row + 1 if ws.max_row else 2
        for col, val in enumerate(values, start=1):
            ws.cell(row=next_row, column=col).value = val

        wb.save(exel_path)
        print(f"[SUCCESS]")
        return True
    
    except Exception as e:
        print(f"[ERROR] Failed ")
        return False